import os  # 导入os库，用来处理文件和文件夹
import cv2  # 导入OpenCV
import openpyxl as xl  # 导入openpyxl模块用来处理Excel文件
from openpyxl.styles import Alignment
import shutil  # 导入shutil用于删除非空文件夹
import numpy as np  # 导入numpy
from PIL import Image  # 从PIL（也就是Python2中的pillow库）中导入Image。PIL(Python Image Library)是python的第三方图像处理库
import tkinter as tk  # 导入tkinter GUI 模块
from tkinter.messagebox import showinfo  # 导入showinfo，用以在窗口中显示说明信息


folderpath = '/Your_Own_Path/'
name_data = folderpath + '/人脸识别文件/'
path = folderpath + '/dataset' # 将人脸图片将要存入的绝对路径保存在变量中
cascadePath = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'   # 级联分类器分类器的实际位置
listpath = name_data + '人脸识别人员信息库.xlsx' # 将人脸识别名单的绝对路径存入变量
ymlpath = folderpath + '/trainer' # YAML文件所在的文件夹绝对路径
yml_file_path = folderpath + '/trainer/trainer.yml'# YAML文件的绝对路径

recognizer = cv2.face.LBPHFaceRecognizer_create()
# 启用局部二进制编码直方图人脸识别器（LBPHFaceRecognizer）并存入变量中
face_detector = cv2.CascadeClassifier(cascadePath)
filename = 'namelist.txt' # 设置人名名单名，并保存入filename变量
font = cv2.FONT_HERSHEY_SIMPLEX # 设置后续将要使用实时识别画面中显示的字体，字体样式为：一般大小的sans-serif字体


def CaptureFaces():  # 此函数用来获取人脸灰度图片，以及用来选择获取方式和保存
    cap = cv2.VideoCapture(0)  # 调用OpenCV内置的VideoCapture()类，参数0意味着使用内置摄像头
    cap.set(3, 640)  # 设置视频宽度
    cap.set(4, 480)  # 设置视频高度
    counter = 0   # 声明一个计数器
    os.makedirs(path + '/' + str(string_name.get())) # 创建以输入的人名命名的文件夹用来储存人脸的灰度图像
    while True:
        ret, img = cap.read()  # 按帧读取图像，如果按帧读取正确，ret返回True，反之返回False。img返回每一帧的图像
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) # 将每一帧的图像从BGR转为灰度
        faces = face_detector.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5, minSize=(20, 20))
        # 使用Haar级联分类器灰度检测人脸，规定缩放比例、最小领域和最小尺寸，
        # detectMultiScale将返回人脸范围在图中的位置（x,y,w,h）

        for(x,y,w,h) in faces:
            cv2.rectangle(img, (x,y), (x+w,y+h), (255,0,0), 2)
            # 以detectMultiScale标定的位置，用蓝色框出人脸，线条粗细为2
            counter += 1
            cv2.imwrite(path + '/' + str(string_name.get()) + "/User." +
                        str(id_none) + '.' + str(string_name.get()) + '.' +
                        str(counter) + ".jpg", gray[y:y + h, x:x + w])
            # 将拍摄好的，已经灰度化的，从原照片切割下来的人脸照片保存
            cv2.imshow('Face sampling...', img)
            # 实时显示拍摄到的，包含人脸的照片（因为在for循环下，如果没有人脸在图像中，则不显示）
        k = cv2.waitKey(100)  # 每一帧图像后，等待100毫秒。如果没有任何键值输入返回-1，有键值输入则返回键值
        if k == 27:  # 如果输入的是ESC键
            break    # 跳出while循环
        elif counter >= 50:  # 计数器不小于50时（50张人脸灰度图像存储完毕）
            break    # 跳出whlie循环
    cap.release()  # 释放拍摄窗口
    cv2.destroyAllWindows()  # 关闭所有OpcnCV打开的窗口

def GetImagesAndLabels(path):  # 此函数用来获取人脸信息样本以及人脸灰度图片对应的id号码，并返回
    imagePaths = []
    i = 0
    for root, dirs, files in os.walk(path):
        i += 1
        if i > 1:
            for fileName in files:
                imagePaths.append(os.path.join(root, fileName))
    # os.walk()用来遍历文件夹下所有文件夹和文件，返回值中，root为每个子文件夹的路径，输出为一个变量，files为所有文件的名称，输出为一个列表
    # os.path.join()用来拼接路径，拼接了path 和 path路径的文件夹中所有文件的名称
    faceSamples, ids = [], []  # 定义空列表，用来存储采集的人脸信息; 定义空列表，用来存储每个人脸图片文件对应的id号码

    for imagePath in imagePaths:
        PIL_img = Image.open(imagePath).convert('L')
        # 用PIL库中的Image打开imagePath下的图像，使用内置函数convert()将图像转换为灰度
        img_numpy = np.array(PIL_img, 'uint8')
        # 将使用Image,open().convert()打开的灰度图像转化为numpy的array数组，数据类型为为八位无符号整型
        id = int(os.path.split(imagePath)[-1].split(".")[1])  # 用于从人脸文件名中，找出用户id号码，并存入id变量
        
        faces = face_detector.detectMultiScale(img_numpy)  
        # 使用Haar级联分类器检测numpy数组中的人脸信息，并返回人脸位置（x,y,w,h）
        for (x,y,w,h) in faces:
            faceSamples.append(img_numpy[y:y+h,x:x+w])  # 读取人脸位置，按顺序存入空列表faceSamples
            ids.append(id)  # 将获取的每个id值，依次存入空列表id
    return faceSamples, ids  # 返回faceSample和id两个列表


def ReturnTraningResult():  # 此函数用来训练人脸特征，同时匹配特征与id，并存入tainner.yml的YAML文件
    faces, ids = GetImagesAndLabels(path)  # 从上一函数中获取返回值
    recognizer.train(faces, np.array(ids)) # 使用LBPH训练人脸样本，并将id存入numpy的数组中
    recognizer.write(yml_file_path)  # 将训练的文件写入


def FaceRecognizer():  # 此函数用来进行人脸识别
    file_open = name_data + '人脸识别人员信息库.xlsx'
    wb_data = xl.load_workbook(file_open)  # 打开'人脸识别人员信息库.xlsx'并读取为当前工作表
    ws_data = wb_data.active  # 读取默认sheet作为当前操作的sheet
    recognizer.read(yml_file_path)  # 读取YAML训练信息文件'人脸识别人员信息库.xlsx'并将
    try:
        recognizer.read(yml_file_path)
    except:
        showinfo(title='提示', message='您还没有进行面部采集！！！')
        window_collect()

    id = 1  # 设id初始值为1
    cap = cv2.VideoCapture(0)  # 打开内置摄像头拍摄
    cap.set(3, 640)  # 设置视频的宽度
    cap.set(4, 480)  # 设置视频的高度

    minW = 0.1*cap.get(3)
    minH = 0.1*cap.get(4)
        # 定义一个最小窗口的尺寸来识别面部
    while True:
        ret, img = cap.read()
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        faces = face_detector.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=6, minSize=(int(minW), int(minH)))
        for(x, y ,w, h) in faces:  # 用OpenCV在人脸范围画出绿色，线条宽度为2的矩形
            id, confidence = recognizer.predict(gray[y:y+h,x:x+w])
            id = id + 2
            # LBPH识别器内置识别函数pridict()，返回相似度和id号码
            if (confidence < 100): # 相似度若小于100，
                cv2.rectangle(img, (x, y), (x+w, y+h), (0, 255, 0), 2)  # 用OpenCV在人脸范围画出绿色，线条宽度为2的矩形
                username = ws_data['B' + str(id)].value  # id 和 callname 列表相互关联，使得可以识别出列表中记录的人名
                confidence = "{0}%".format(round(100 - confidence))
                cv2.putText(img, str(confidence), (x + 5, y + h - 5), font, 1, (255, 255, 0), 1)
                # 在识别画面中显示相似度
                cv2.putText(img, str(username), (x+5, y-5), font, 1, (255, 255, 255), 2)  # 在识别画面中显示人名
            else:
                username = "Unknown"
                cv2.rectangle(img, (x, y), (x+w, y+h), (0, 0, 255), 2)  # 用OpenCV在人脸范围画出红色，线条宽度为2的矩形
                confidence = "{0}%".format(round(100 - confidence))
                cv2.putText(img, str(username), (x+5, y-5), font, 1, (0, 0, 255), 2)  #在识别画面中显示unknown

        cv2.imshow('Recognizing...  Press Key "ESC" to Exit.',img)  # 显示实时识别画面
        k = cv2.waitKey(10)  # 等待操作退出while循环
        if k == 27:
            # 通过opencv显示的窗口有一个默认属性autosize，显然当窗口关闭这个属性自然不存在了，所以在cv2.imshow之后添加
            # 判断窗口属性值是否是autosize，如果不是就退出循环，销毁窗口。可以使用鼠标点击关闭窗口
            break  # 退出while循环
    cap.release()
    cv2.destroyAllWindows()
'''******************************************************************************************************************'''
#                                   以下是数据部分代码（使用xlsx文件读写实现）                                                                       
'''******************************************************************************************************************'''
def set_data_xl():  # 建立人脸识别的人员信息库
    wb = xl.Workbook()
    ws = wb.active
    # ————————设置列名——————————
    ws['A2'] = '编号'  # 设置 A2 单元格为'编号'
    ws['B2'] = '姓名'
    ws['C2'] = '人名数量'
    ws['A1'] = '人员信息库'
    ws.merge_cells('A1:C1')  # 合并单元格从 A1 到 C1，则合并后的单元格将显示 A1 的内容
    # ————————设置列宽——————————
    ws.column_dimensions['B'].width = 12  # 设置 B 列的列宽
    ws.column_dimensions['C'].width = 16
    # ————————逐列居中——————————
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')  # 将 A1 单元格设置格式为垂直居中和水平居中
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    wb.save(name_data + '人脸识别人员信息库.xlsx')  # 将当前操作保存为新的'人脸识别人员信息库.xlsx'文件


def detect_name_num():  # 检测姓名列的人名数量
    file_open = name_data + '人脸识别人员信息库.xlsx'
    wb_data = xl.load_workbook(file_open)
    ws_data = wb_data.active
    all_rows = ws_data.max_row  # 取总行数赋值给 all_rows 变量
    name_rows = 0  # 设置人名列的计数变量
    none = 0  # 设置姓名列（B 列）空行计数变量
    blank = ws_data['B' + str(name_rows + 3)].value
    while name_rows == all_rows - 2:  # 当人名行数（不包含中间的空行） == 总函数-2时证明已经查找完
        if blank is not None:  # 当 blank 变量（B 列的任意单元格）有内容时
            blank = ws_data['B' + str(name_rows + 3)].value  # 再次定义，否则会出现错误
            name_rows += 1  # 计数变量自增一
        elif blank is None:  # 当 blank 是空的时候（没有内容）
            blank = ws_data['B' + str(name_rows + 3)].value  # 再次定义，否则会出现错误
            name_rows += 1  # 计数变量自增一
            none += 1  # 空行计数变量自增一
    name_counter = name_rows - none  # 设置人名的数量（不含人名列中的空行）
    ws_data['C3'] = name_counter  # 人名数量将显示在 C3 单元格
    wb_data.save(name_data + '人脸识别人员信息库.xlsx')


def insert_name_on_blank():  # 用于实现置入人名于 B列 自上而下第一个空单元格
    file_open = name_data + '人脸识别人员信息库.xlsx'
    wb = xl.load_workbook(file_open)
    ws = wb.active
    check = 1  # 初始姓名行数，用于检测非空行的计数器
    blank = ws['B' + str(check+2)].value  # 将 B列 任意单元格的内容赋值给 lank 变量
    count = ws['C3'].value  # 先读取由detect_name_num()函数存在文件中的人名数量，赋值给 count 变量
    count += 1  # 因调取本函数时，肯定会输入人名，所以，count 一定要自增一
    ws['C3'] = count  # 将自增后的 count 变量存回 C3 单元格
    if blank is not None:  # 若 blank 变量不是空（所指的单元格有内容）
        while blank is not None:  # 同上条件的while
            blank = ws['B' + str(check+2)].value  # 重新赋值，以免出错，check+2 是为了直接操作第三行的数据，防止误操作表头的两行
            check += 1  # 非空行数加一，亦即人名数加一
        ws['B' + str(check+1)] = string_name.get() #
        ws['A' + str(check+1)] = str(check-1)
    elif blank is None:
        ws['A' + str(check+2)] = str(check)
        ws['B' + str(check+2)] = string_name.get()
    wb.save(name_data + '人脸识别人员信息库.xlsx')
'''**************************************************************************************************************'''
#                                             以下是GUI代码
'''**************************************************************************************************************'''
root = tk.Tk()
root.title("人脸识别系统")


# ――――――――――――――――——————————根窗口―――――――――――――――――――――――――――——
def root_window():  # 根窗口
    tk.Label(root, text="人脸识别系统", font=("华文行楷", 30)).grid(columnspan=3, sticky="ew")  # 定义标题
    tk.Label(root).grid(row=1, column=0)  # 三个标签用来占位
    tk.Label(root).grid(row=2, column=1)
    tk.Label(root).grid(row=2, column=1)
    # 设置5个功能按钮，root为指向的窗口，font为显示的字体与大小，command为按钮的命令,grid()为位置
    tk.Button(root, text="面部采集", font=("宋体", 20), command=window_collect).grid(row=2, column=0)
    tk.Button(root, text="人脸识别", font=("宋体", 20), command=window_recognize).grid(row=2, column=2)
    tk.Button(root, text="使用说明", font=("宋体", 20), command=window_illustration).grid(row=3, column=0)
    tk.Button(root, text="删除功能", font=("宋体", 20), command=window_delete).grid(row=3, column=2)
    tk.Button(root, text='程序初始化', font=("宋体", 20), command=window_initialize).grid(row=4, column=0)
    tk.Label(root, text="提示：开始前请先点击阅读使用说明！", font=("楷体", 25)).grid(row=5, column=0, columnspan=3, sticky="ew")
    # 使得每一列拥有跟随拉伸的性质
    root.grid_columnconfigure(0, weight=1)
    root.grid_columnconfigure(1, weight=1)
    root.grid_columnconfigure(2, weight=1)
    root.grid_columnconfigure(3, weight=1)
    root.geometry("800x600")  # 设置窗口默认大小为800x600像素
    isexist = os.path.exists(folderpath)
    if isexist == False:  # 判断程序所需的数据文件夹是否存在，如果不存在，自动进行初始化
        program_initialize()
    root.mainloop()


# ―――――――——————————————二级窗口--面部采集―――――――――――――――――――――――――――
def window_collect():
    global win_colt
    file_open = name_data + '人脸识别人员信息库.xlsx'
    wb = xl.load_workbook(file_open)
    ws = wb.active
    check_none = 1  # 初始姓名行数，用于检测空行
    blank = ws['B' + str(check_none + 2)].value
    global id_none
    if blank is not None:  # 为显示提示中的id号码而写的
        while blank is not None:
            blank = ws['B' + str(check_none + 2)].value
            check_none += 1
        id_none = check_none-1
    elif blank is None:
        id_none = check_none
    win_colt = tk.Tk()
    win_colt.title("面部采集")
    win_colt.geometry("800x600")

    tk.Label(win_colt).grid(row=0, column=0)
    tip = tk.LabelFrame(win_colt, text="使用提示：")
    tip.grid(row=2, column=0)
    message_text = '请记下分配给被识别人的ID号码，并输入被识别人姓名（拼音全写）点击“采集开始”后，按钮不会立即弹起，加载采集画面需要几秒钟时间，请耐心等待。采集时，' \
                   '请保持面部正面被拍摄，并保持几秒，待采集窗口自动关闭，请关闭"面部采集窗口"，再进行其他操作。每次打开本窗口，仅可输入一人的信息！！！'
    tk.Message(tip, text=message_text, font=("宋体", 15), bg="khaki").grid(row=3, column=0)

    global string_name
    string_name = tk.StringVar(master=win_colt)
    tip_3 = tk.LabelFrame(win_colt, text='此人分配到的ID为' + str(id_none))
    tip_3.grid(row=2, column=1, padx=30)
    tk.Label(tip_3, text="请输入被识别人姓名：").grid(row=3, column=1)
    tk.Entry(tip_3, textvariable=string_name).grid(row=4, column=1)
    tk.Button(tip_3, text="采集开始", font=("宋体", 20), command=start_collect).grid(row=4, column=2)

    win_colt.mainloop()


# ――――――――――――――――——————二级窗口--人脸识别―――――――――――――――――――――――――――——————
def window_recognize():  # 人脸识别窗口
    global win_rcn
    win_rcn = tk.Tk()
    win_rcn.title("人脸识别")
    win_rcn.geometry("600x470")

    tk.Label(win_rcn).grid(row=0, column=0)
    tk.Button(win_rcn, text="开始人脸识别", font=("宋体", 20), command=start_recognize).grid(row=1, column=0, padx=25)
    # ――――――――――――――――——————————按钮提示――――――――――――――――――――――――――――—————————
    tip_1 = tk.LabelFrame(win_rcn, text="使用提示：")
    tip_1.grid(row=2, column=0)
    message_text = "点击按钮后，按钮不会立即弹起，加载识别画面需要几秒钟时间，请耐心等待。\n请按“ESC”键来退出窗口。"
    tk.Message(tip_1, text=message_text, font=("宋体", 15), bg="khaki").grid(row=3, column=0)
    # 当没有任何人员信息时，进行异常的处理，反馈问题并弹出面部采集窗口
    try:
        recognizer.read(yml_file_path)
    except:
        showinfo(title='提示', message='您还没有进行面部采集！！！')
        window_collect()
    win_rcn.mainloop()

# ――――――――――――――――————————————二级窗口--删除功能————―――――――――――――――――――――――――――
def window_delete():  # 删除功能窗口
    global win_del  # 全局化变量
    global sgl_selection
    win_del = tk.Tk()
    sgl_selection = tk.IntVar(master=win_del)
    win_del.title("删除功能")
    win_del.geometry("600x470")
    tk.Label(win_del, text='请选择删除方式：').grid(row=1, column=0)
    tk.Radiobutton(win_del, text="删除全部信息", variable=sgl_selection, value=0).grid(row=2, column=0)
    tk.Radiobutton(win_del, text="删除某人信息", variable=sgl_selection, value=1).grid(row=3, column=0)
    tk.Button(win_del, text="确定", command=start_delete).grid(row=3, column=1)
    # ――――――――――――――――————————使用提示――――――――――――――――――――――――――――——
    tip_2 = tk.LabelFrame(win_del, text="使用提示：")
    tip_2.grid(row=0, column=2)
    message_text = "选择“删除全部信息”将会删除包括已录入的图像和人名信息；选择“删除某人信息“" \
                   "将会删除某人已经录入的人名录入和此人的所有照片。若程序数据文件不存在，将直接初始化并转入面部采集窗口"
    tk.Message(tip_2, text=message_text, bg="khaki").grid(row=2, column=2)

    win_del.mainloop()


# ――――――――――――――――————————二级窗口--使用说明―――――――――――――――――――――――――――
def window_illustration():  # 使用说明弹窗
    showinfo(title='使用说明', message='欢迎使用本系统，请先进行面部采集再进行人脸识别。删除功能提供两种选项：“删除全部'
                                   '信息”和“删除某人信息”供您选择。\n初次使用，请先点击程序初始化！\n作者')


# ――――――――――――――――————————二级窗口--初始化―――――――――――――――――――――――――――—
def window_initialize():  
    win_ini = tk.Tk()
    win_ini.title('初始化')
    
    def handle_confirm():
        win_ini.destroy()  # 关闭初始化确定窗口
        program_initialize()  # 执行初始化逻辑
        tk.messagebox.showinfo('通知', '初始化已经完成！')  # 弹出完成通知
    
    tk.Label(win_ini, text='您确定进行初始化吗? ').pack()
    tk.Button(win_ini, text='确定', command=handle_confirm).pack()  # 绑定新处理函数
    win_ini.mainloop()


# ――――――――――――――――——————二级窗口中的按钮动作――――――――――――――――――――――――――
def start_collect():  # 开始面部采集按钮触发的函数
    insert_name_on_blank()
    CaptureFaces()
    GetImagesAndLabels(path)
    ReturnTraningResult()
    showinfo(title='提示', message='采集完成！请关闭‘面部采集窗口’')


def start_recognize():
    FaceRecognizer()


def start_delete():
    if sgl_selection.get() == 1:  # 删除单个人
        global del_name
        del_name = tk.StringVar(master=win_del)
        tk.Label(win_del).grid(row=0, column=2)
        tip_1 = tk.LabelFrame(win_del, text="请输入要删除的人员姓名：")
        tip_1.grid(row=1, column=2)
        tk.Button(tip_1, text='确定', command=sgl_delete_command).grid(row=3, column=2)

    elif sgl_selection.get() == 0:  # 清空数据
        shutil.rmtree(name_data)  # 先删除整个文件夹 # 重建照片文件夹
        os.makedirs(name_data)

def sgl_delete_command():
    file_open = name_data + '人脸识别人员信息库.xlsx'
    wb_data = xl.load_workbook(file_open)
    ws_data = wb_data.active
    name_rows = 1  # 人名初始行数
    if not del_name.get() is None:
        shutil.rmtree(path + '/' + str(del_name.get()))  # 删除含有照片的人名文件夹
    while True:
        finder = ws_data['B' + str(name_rows+2)].value
        name_rows += 1
        if finder == del_name.get():
            info = '已删除姓名为： ' + finder + ' 的人员信息'
            showinfo(title="提示", message=info)
            ws_data['B' + str(name_rows + 1)].value = None
            ct = ws_data['C3'].value
            ct = ct - 1
            ws_data['C3'] = ct
            wb_data.save(name_data + '人脸识别人员信息库.xlsx')
            break


def program_initialize():
    is_exist_1 = os.path.exists(path)
    is_exist_2 = os.path.exists(ymlpath)
    is_exist_3 = os.path.exists(name_data)
    if is_exist_1 == True:
        shutil.rmtree(path)
        os.makedirs(path)
    else:
        os.makedirs(path)
    if is_exist_2 == True:
        shutil.rmtree(ymlpath)
        os.makedirs(ymlpath)
    else:
        os.makedirs(ymlpath)
    if is_exist_3 == True:
        shutil.rmtree(name_data)
        os.makedirs(name_data)
    else:
        os.makedirs(name_data)
    set_data_xl()
    detect_name_num()

root_window()